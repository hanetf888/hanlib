import os
import datetime
from datetime import timedelta
import win32com.client as client
import time
from openpyxl.styles import Border, Side
from pathlib import Path
import inspect
import zipfile
import xlrd
import csv
import logging

logger = logging.getLogger(__name__)




def is_business_day(date_obj):
    """
    Checks if a given date is a business day (Monday-Friday).
    Does not account for public holidays.
    """
    # Monday is 0, Sunday is 6
    return date_obj.weekday() < 5


def getlastbusinessday(run_date, days_offset):
    """
    rundate must be a datetime.date
    days_offset must be an integer

    Returns the last business date according to the offset.
    days_offset = -1 will return the previous business day
    days_offset = -2 will return the business day 2 days ago

    function returns -1 if there is an error

    """

    if not isinstance(run_date, datetime.date):
        print("utils.getlastbusinessday: run_date passed is not a datetime.date")
        return -1

    if not isinstance(days_offset, int):
        print("utils.getlastbusinessday: days_offset passed is not an int")
        return -1

    run_date_with_offset = run_date + timedelta(days=days_offset)

    # if the run_date with the offset applied is a weekend then we keep
    # subtracting days until it is a weekday.
    while run_date_with_offset.weekday() > 4:
        if days_offset == -2:
            run_date_with_offset = run_date_with_offset + timedelta(days=days_offset)
        else:
            run_date_with_offset = run_date_with_offset + timedelta(
                days=days_offset - 1
            )
    return run_date_with_offset


def get_busday_x_days_from_date(start_date, x_days):
    """
    Returns a date that is 'business_days' from 'start_date', excluding weekends and optional holidays.

    :param start_date: (datetime or str) The starting date (format: 'YYYY-MM-DD' if string)
    :param business_days: (int) Number of business days to add (can be negative for past dates)
    :param holidays: (list of str) Optional list of holiday dates (format: 'YYYY-MM-DD')
    :return: (datetime) The resulting business date
    """
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, "%Y-%m-%d")

    # holidays = set(datetime.strptime(h, "%Y-%m-%d").date() for h in (holidays or []))

    current_date = start_date
    step = 1 if x_days > 0 else -1  # Forward or backward

    while x_days != 0:
        current_date += timedelta(days=step)
        # if current_date.weekday() < 5 and current_date not in holidays:  # Monday-Friday and not a holiday
        if current_date.weekday() < 5:  # Monday-Friday
            x_days -= step

    return current_date


def get_first_business_day_of_month(year, month):
    """
    Calculates the first business day of a given month.
    """
    day = 1
    while True:
        try:
            current_date = datetime.date(year, month, day)
            if is_business_day(current_date):
                return current_date
            day += 1
        except ValueError:
            # If we try to create a date like Feb 30, it will raise ValueError.
            # This means we've gone past the end of the month, which shouldn't happen
            # if the month has at least one business day (which it always will).
            raise Exception(
                f"Could not find a business day in {month}/{year}. This should not happen."
            )


def is_weekend(date):
    """
    Checks if the given date is a weekend (Saturday or Sunday).

    :param date: (datetime.date) The date to check.
    :return: (bool) True if it's a weekend, False otherwise.
    """
    return date.weekday() >= 5  # 5 = Saturday, 6 = Sunday


def mailsubjectexists(subject, folder, email):
    outlook = client.Dispatch("outlook.application")
    mapi = outlook.GetNamespace("MAPI")

    messages = mapi.Folders(email).Folders(folder).Items

    messages = messages.Restrict(f"[Subject] = '{subject}'")

    if len(messages) == 0:
        # If the length of the list is zero then the email is not found - return false
        return False
    else:
        # If the length of the list is> 0 then the email has been found - return true
        return True


def setborder(ws, startrow, startcol, endrow, endcol):
    thin = Side(border_style="thin", color="000000")

    if endrow < startrow:
        raise ValueError("Start end row is before start row")
    if endcol < startcol:
        raise ValueError("Start end column is before start column")

    for col in range(startcol, endcol + 1):
        ws.cell(startrow, col).border = Border(top=thin)

    for col in range(startcol, endcol + 1):
        ws.cell(endrow, col).border = Border(bottom=thin)

    for rw in range(startrow, endrow):
        ws.cell(rw, startcol).border = Border(left=thin)
        ws.cell(rw, endcol).border = Border(right=thin)

    # Set corners
    ws.cell(startrow, startcol).border = Border(top=thin, left=thin)
    ws.cell(startrow, endcol).border = Border(top=thin, right=thin)
    ws.cell(endrow, startcol).border = Border(bottom=thin, left=thin)
    ws.cell(endrow, endcol).border = Border(bottom=thin, right=thin)


# getlastmonthend , months must be a positive number.  3 measn it gets 3 monthends ago
def getlastmonthend(curdate, months):

    returndate = curdate

    for i in range(months):
        returndate = datetime.date(
            returndate.year, returndate.month, 1
        ) - datetime.timedelta(days=1)

    return returndate


def getlastmonthendbd(curdate, months):
    returndate = curdate

    for i in range(months):
        returndate = datetime.date(
            returndate.year, returndate.month, 1
        ) - datetime.timedelta(days=1)

    return getlastbusinessday(returndate, 0)




def doNothing():
    pass


def getFinalTermsRegulator(filename: str):
    regulators = ["FCA", "CBI", "SFSA"]
    for regulator in regulators:
        if filename.lower().find(regulator.lower()) >= 0:
            return regulator


def findFile(
    folder: str,
    keyword: str,
    extension: str,
    how: str = "latest",
    returns: str = "filepath",
) -> str:
    """Provide a folder, keyword, and extension, along with how ('latest' or 'earliest') and what to return ('filepath' or 'date')
    and this returns the (latest or earliest) (filepath or saved date) with that keyword and extension in the specified folder, as a (string or datetime.date object).
    """

    files = os.listdir(folder)

    relevant_files = []
    relevant_dates = []

    for filename in files:

        filepath = folder + "\\" + filename
        created_date = datetime.datetime.fromtimestamp(os.path.getmtime(filepath))

        if filepath.find(keyword) >= 0 and filepath.find(extension) >= 0:

            relevant_files.append(filepath)
            relevant_dates.append(created_date)

    relevant_file_date = max(relevant_dates) if how == "latest" else min(relevant_dates)
    relevant_index = relevant_dates.index(relevant_file_date)
    relevant_filename = relevant_files[relevant_index]

    return relevant_filename if returns == "filepath" else relevant_file_date.date()


def castAsString(data):
    try:
        return str(data)
    except Exception:
        return data


def toDate(data, to="date"):
    if to == "date":
        return datetime.datetime.date(datetime.datetime.strptime(data, "%Y%m%d"))
    else:
        return datetime.datetime.strftime(data, "%Y%m%d")


def divByMillion(data):

    return data / 1000000


def percent(data):
    return data / 100


def current_function_name():
    return inspect.currentframe().f_code.co_name


class FolderNotFound(Exception):
    """Custom exception for when folder is not found."""

    pass


def zip_and_clean_folder_with_exclusions(curdate, folder_path, excludelist):
    # Extract the folder name from the folder path
    foldername = os.path.basename(os.path.abspath(folder_path))
    lastbusdate = get_busday_x_days_from_date(curdate, -1)

    # Check if the folder exists
    if not os.path.exists(folder_path):
        raise FolderNotFound(f"The folder '{folder_path}' does not exist.")

    # Path for the zip file
    zip_file_path = os.path.join(folder_path, f"{foldername}.zip")

    # Gather a list of all files in the folder excluding the excludelist and the ZIP file itself
    files_to_zip = [
        f
        for f in os.listdir(folder_path)
        if os.path.isfile(os.path.join(folder_path, f))
        and f not in excludelist  # Exclude files in the excludelist
        and f != f"{foldername}.zip"  # Exclude the target folder ZIP file
        and curdate.strftime("%Y%m%d")
        not in f  # Exclude files with the current date in the filename
        and curdate.strftime("%d.%m.%Y")
        not in f  # Exclude files with the current date in the filename
        and curdate.strftime("%d_%m_%Y")
        not in f  # Exclude files with the current date in the filename
        and curdate.strftime("%d%m")
        not in f  # Exclude files with the current date in the filename
        and curdate.strftime("%d %m %Y")
        not in f  # Exclude files with the current date in the filename
        and lastbusdate.strftime("%Y%m%d")
        not in f  # Exclude files with the current date in the filename
        and lastbusdate.strftime("%d.%m.%Y")
        not in f  # Exclude files with the current date in the filename
        and lastbusdate.strftime("%d_%m_%Y")
        not in f  # Exclude files with the current date in the filename
        and lastbusdate.strftime("%d%m")
        not in f  # Exclude files with the current date in the filename
        and lastbusdate.strftime("%d %m %Y")
        not in f  # Exclude files with the current date in the filename
    ]

    # If the zip file does not exist, create it
    if not os.path.exists(zip_file_path):
        with zipfile.ZipFile(zip_file_path, "w") as zf:
            for file in files_to_zip:
                zf.write(os.path.join(folder_path, file), arcname=file)
        logger.info(f"Created new zip file: {zip_file_path}")
    else:
        # If the zip file exists, add any files not already in it and not in the excludelist
        with zipfile.ZipFile(zip_file_path, "a") as zf:
            existing_files = set(zf.namelist())  # Files already in the zip
            for file in files_to_zip:
                if file not in existing_files:
                    zf.write(os.path.join(folder_path, file), arcname=file)
                    file_path = os.path.join(folder_path, file)

                    logger.info(
                        f"Zipped file {file_path} to existing zip file {zip_file_path}"
                    )

                    # If the file is in the zip file, not on the exclude list and exits
                    # in the folder, delete it
                    if (
                        file not in excludelist
                        and file in zf.namelist()
                        and os.path.exists(file_path)
                    ):
                        os.remove(file_path)
                        logger.info(
                            f"Deleted file: {file_path} as in zipfile {zip_file_path}"
                        )

        logger.info(f"Updated existing zip file: {zip_file_path}")

    # # Delete files that are now zipped, except files listed in the
    # with zipfile.ZipFile(zip_file_path, 'r') as zf:
    #     zipped_files = zf.namelist()  # Files present in the zip
    #     for file in zipped_files:
    #         if file not in excludelist:  # Skip files in the excludelist
    #             file_path = os.path.join(folder_path, file)





def convert_xls_to_csv_with_formats(xls_filename, csv_filename):
    # Open the .xls file
    workbook = xlrd.open_workbook(xls_filename)
    sheet = workbook.sheet_by_index(
        0
    )  # Use the first sheet (can be adjusted dynamically)

    # Open the .csv file for writing
    with open(csv_filename, "w", newline="") as csv_file:
        csv_writer = csv.writer(csv_file)

        # Loop through rows in the Excel sheet
        for row_idx in range(sheet.nrows):
            row = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                cell_type = cell.ctype  # Get the type of the cell
                cell_value = cell.value

                if cell_type == xlrd.XL_CELL_DATE:  # Handle date types
                    # Convert Excel serial date to a Python datetime object
                    date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                    cell_value = datetime.datetime(*date_tuple).strftime(
                        "%Y-%m-%d"
                    )  # Format the date as YYYY-MM-DD
                elif cell_type == xlrd.XL_CELL_NUMBER:  # Handle numbers
                    cell_value = (
                        "%.15g" % cell_value
                    )  # Handle large numbers and preserve precision

                row.append(cell_value)  # Append the formatted cell value to the row
            logger.info(f"{row}")
            csv_writer.writerow(row)  # Write the row to the CSV

    logger.info(
        f"Converted '{xls_filename}' to '{csv_filename}' with formats preserved!"
    )

def excel_open_force_recalc(filename, excel_tab, cell_update):

    try: 
        excel = client.DispatchEx('Excel.Application')
        # Display the excel application
        excel.Visible = True

        #turn off all dialog boxes and notificatioins
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.AutomationSecurity = 1  # msoAutomationSecurityLow

        # 1. open the file
        workbook = excel.Workbooks.Open(filename, UpdateLinks=0, ReadOnly=False, IgnoreReadOnlyRecommended=True)
        logger.info(rf'Workbook "{filename}" opened in Excel.')
        
        # 2. move to the tab to update 
        sheet = workbook.Sheets(excel_tab)
        sheet.Activate()
        logger.info(f"Activated sheet: {excel_tab}")
        time.sleep(5)  # give Excel time to move to sheet

        # 3. Perform an Action (e.g., updating cell to force a recalculation and refresh cache)
        sheet.Range(cell_update).Value = " "   
        logger.info(f"Updated cell {cell_update} on tab {excel_tab} value to force recalculation.")

        # 3.1 Force Excel to recalculate all formulas as well
        workbook.Application.CalculateFull()
        logger.info("Forced full recalculation of all formulas in workbook.")

        # Give Excel a moment to recalculate (especially useful for complex workbooks)
        time.sleep(5) 
        
        # 4. Save the Workbook
        workbook.Save()
        logger.info(f"Workbook {filename} saved successfully.")

    except Exception as e:
        logger.error(f"An error occurred while processing the Excel file: {filename} {e}")
        
    finally:    
        workbook.Close(SaveChanges=False)
        excel.Quit()    
        logger.info(f"Excel closed successfully '{filename}'.")
        time.sleep(5) 

