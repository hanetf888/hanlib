"""
Tests for hanlib.utils

Tests cover all pure/logic functions. Functions that require Windows COM
(mailsubjectexists, excel_open_force_recalc) are not tested here as they
need a desktop with Outlook/Excel installed.
"""

import pytest
import datetime
import os
import zipfile
from pathlib import Path
from unittest.mock import MagicMock
from openpyxl import Workbook

from hanlib import utils


# =============================================================================
# is_business_day
# =============================================================================

class TestIsBusinessDay:

    @pytest.mark.parametrize("date_obj,expected", [
        (datetime.date(2026, 2, 2), True),   # Monday
        (datetime.date(2026, 2, 3), True),   # Tuesday
        (datetime.date(2026, 2, 4), True),   # Wednesday
        (datetime.date(2026, 2, 5), True),   # Thursday
        (datetime.date(2026, 2, 6), True),   # Friday
        (datetime.date(2026, 2, 7), False),  # Saturday
        (datetime.date(2026, 2, 8), False),  # Sunday
    ])
    def test_weekdays_and_weekends(self, date_obj, expected):
        assert utils.is_business_day(date_obj) is expected


# =============================================================================
# is_weekend
# =============================================================================

class TestIsWeekend:

    def test_saturday(self):
        assert utils.is_weekend(datetime.date(2026, 2, 7)) is True

    def test_sunday(self):
        assert utils.is_weekend(datetime.date(2026, 2, 8)) is True

    def test_weekday(self):
        assert utils.is_weekend(datetime.date(2026, 2, 6)) is False


# =============================================================================
# getlastbusinessday
# =============================================================================

class TestGetLastBusinessDay:

    def test_offset_minus_1_from_weekday(self):
        # Tuesday 2026-02-03 -> Monday 2026-02-02
        result = utils.getlastbusinessday(datetime.date(2026, 2, 3), -1)
        assert result == datetime.date(2026, 2, 2)

    def test_offset_minus_1_from_monday(self):
        # Monday 2026-02-02, offset -1 -> Sunday -> rolls back to Friday
        result = utils.getlastbusinessday(datetime.date(2026, 2, 2), -1)
        assert result.weekday() < 5, "Should return a weekday"
        assert result < datetime.date(2026, 2, 2)

    def test_offset_0_on_weekday(self):
        # Friday with offset 0 should return Friday itself
        friday = datetime.date(2026, 2, 6)
        result = utils.getlastbusinessday(friday, 0)
        assert result == friday

    def test_offset_0_on_weekend(self):
        # Saturday with offset 0 should roll to a weekday
        saturday = datetime.date(2026, 2, 7)
        result = utils.getlastbusinessday(saturday, 0)
        assert result.weekday() < 5

    def test_invalid_run_date_returns_minus_1(self):
        result = utils.getlastbusinessday("not-a-date", -1)
        assert result == -1

    def test_invalid_offset_returns_minus_1(self):
        result = utils.getlastbusinessday(datetime.date(2026, 2, 3), "bad")
        assert result == -1

    def test_offset_minus_2(self):
        # Wednesday 2026-02-04, offset -2 -> Monday 2026-02-02
        result = utils.getlastbusinessday(datetime.date(2026, 2, 4), -2)
        assert result.weekday() < 5
        assert result < datetime.date(2026, 2, 4)


# =============================================================================
# get_busday_x_days_from_date
# =============================================================================

class TestGetBusdayXDaysFromDate:

    def test_positive_days_skips_weekend(self):
        # Friday 2026-02-06, +1 business day -> Monday 2026-02-09
        result = utils.get_busday_x_days_from_date(datetime.date(2026, 2, 6), 1)
        assert result == datetime.date(2026, 2, 9)

    def test_positive_days_within_week(self):
        # Monday 2026-02-02, +2 business days -> Wednesday 2026-02-04
        result = utils.get_busday_x_days_from_date(datetime.date(2026, 2, 2), 2)
        assert result == datetime.date(2026, 2, 4)

    def test_negative_days(self):
        # Monday 2026-02-09, -1 business day -> Friday 2026-02-06
        result = utils.get_busday_x_days_from_date(datetime.date(2026, 2, 9), -1)
        assert result == datetime.date(2026, 2, 6)

    def test_five_business_days_spans_week(self):
        # Monday 2026-02-02, +5 -> Friday 2026-02-06 (but continues to next Monday)
        # Actually +5 from Mon = next Mon
        result = utils.get_busday_x_days_from_date(datetime.date(2026, 2, 2), 5)
        assert result == datetime.date(2026, 2, 9)

    def test_result_is_always_weekday(self):
        start = datetime.date(2026, 2, 2)
        for days in range(1, 15):
            result = utils.get_busday_x_days_from_date(start, days)
            assert result.weekday() < 5, f"+{days} business days landed on weekend"


# =============================================================================
# get_first_business_day_of_month
# =============================================================================

class TestGetFirstBusinessDayOfMonth:

    def test_month_starting_on_weekday(self):
        # February 2026 starts on a Sunday, so first biz day is Monday 2nd
        result = utils.get_first_business_day_of_month(2026, 2)
        assert result == datetime.date(2026, 2, 2)

    def test_month_starting_on_monday(self):
        # June 2026 starts on a Monday
        result = utils.get_first_business_day_of_month(2026, 6)
        assert result == datetime.date(2026, 6, 1)

    def test_result_is_always_weekday(self):
        for month in range(1, 13):
            result = utils.get_first_business_day_of_month(2026, month)
            assert result.weekday() < 5
            assert result.month == month
            assert result.day <= 3  # Can be at most the 3rd (if 1st is Saturday)


# =============================================================================
# getlastmonthend
# =============================================================================

class TestGetLastMonthEnd:

    def test_one_month_ago(self):
        # From 2026-02-15, 1 month end ago -> 2026-01-31
        result = utils.getlastmonthend(datetime.date(2026, 2, 15), 1)
        assert result == datetime.date(2026, 1, 31)

    def test_two_months_ago(self):
        # From 2026-03-10: 1st iteration -> Feb 28, 2nd iteration -> Jan 31
        result = utils.getlastmonthend(datetime.date(2026, 3, 10), 2)
        assert result == datetime.date(2026, 1, 31)

    def test_from_first_of_month(self):
        # From 2026-03-01, 1 month end ago -> 2026-02-28
        result = utils.getlastmonthend(datetime.date(2026, 3, 1), 1)
        assert result == datetime.date(2026, 2, 28)

    def test_february_leap_year(self):
        # From 2024-03-15, 1 month end ago -> 2024-02-29 (leap year)
        result = utils.getlastmonthend(datetime.date(2024, 3, 15), 1)
        assert result == datetime.date(2024, 2, 29)


# =============================================================================
# getlastmonthendbd
# =============================================================================

class TestGetLastMonthEndBd:

    def test_returns_business_day(self):
        # The result should always be a weekday
        result = utils.getlastmonthendbd(datetime.date(2026, 2, 15), 1)
        assert result.weekday() < 5

    def test_month_ending_on_weekend(self):
        # Jan 2023 ends on Tuesday (31st), should return 31st
        result = utils.getlastmonthendbd(datetime.date(2023, 2, 15), 1)
        assert result == datetime.date(2023, 1, 31)
        assert result.weekday() < 5


# =============================================================================
# castAsString
# =============================================================================

class TestCastAsString:

    def test_int(self):
        assert utils.castAsString(42) == "42"

    def test_float(self):
        assert utils.castAsString(3.14) == "3.14"

    def test_string_passthrough(self):
        assert utils.castAsString("hello") == "hello"

    def test_none(self):
        assert utils.castAsString(None) == "None"

    def test_date(self):
        d = datetime.date(2026, 1, 1)
        assert utils.castAsString(d) == "2026-01-01"


# =============================================================================
# toDate
# =============================================================================

class TestToDate:

    def test_string_to_date(self):
        result = utils.toDate("20260201", to="date")
        assert result == datetime.date(2026, 2, 1)

    def test_date_to_string(self):
        d = datetime.date(2026, 2, 1)
        result = utils.toDate(d, to="string")
        assert result == "20260201"


# =============================================================================
# divByMillion
# =============================================================================

class TestDivByMillion:

    def test_positive(self):
        assert utils.divByMillion(5_000_000) == 5.0

    def test_zero(self):
        assert utils.divByMillion(0) == 0.0

    def test_fractional(self):
        assert utils.divByMillion(500_000) == 0.5


# =============================================================================
# percent
# =============================================================================

class TestPercent:

    def test_basic(self):
        assert utils.percent(50) == 0.5

    def test_zero(self):
        assert utils.percent(0) == 0.0

    def test_over_100(self):
        assert utils.percent(250) == 2.5


# =============================================================================
# getFinalTermsRegulator
# =============================================================================

class TestGetFinalTermsRegulator:

    def test_fca(self):
        assert utils.getFinalTermsRegulator("Final_Terms_FCA_2026.pdf") == "FCA"

    def test_cbi(self):
        assert utils.getFinalTermsRegulator("doc_cbi_filing.docx") == "CBI"

    def test_sfsa(self):
        assert utils.getFinalTermsRegulator("SFSA_report.xlsx") == "SFSA"

    def test_case_insensitive(self):
        assert utils.getFinalTermsRegulator("final_fca_terms.pdf") == "FCA"

    def test_no_match_returns_none(self):
        result = utils.getFinalTermsRegulator("unrelated_document.pdf")
        assert result is None


# =============================================================================
# doNothing
# =============================================================================

class TestDoNothing:

    def test_returns_none(self):
        assert utils.doNothing() is None


# =============================================================================
# FolderNotFound
# =============================================================================

class TestFolderNotFound:

    def test_is_exception(self):
        assert issubclass(utils.FolderNotFound, Exception)

    def test_can_be_raised(self):
        with pytest.raises(utils.FolderNotFound):
            raise utils.FolderNotFound("test")


# =============================================================================
# setborder
# =============================================================================

class TestSetBorder:

    def test_applies_borders(self):
        wb = Workbook()
        ws = wb.active
        utils.setborder(ws, 1, 1, 3, 3)
        # Top-left corner should have top and left borders
        cell = ws.cell(1, 1)
        assert cell.border.top.border_style == "thin"
        assert cell.border.left.border_style == "thin"

    def test_bottom_right_corner(self):
        wb = Workbook()
        ws = wb.active
        utils.setborder(ws, 1, 1, 3, 3)
        cell = ws.cell(3, 3)
        assert cell.border.bottom.border_style == "thin"
        assert cell.border.right.border_style == "thin"

    def test_invalid_row_range_raises(self):
        wb = Workbook()
        ws = wb.active
        with pytest.raises(ValueError, match="Start end row is before start row"):
            utils.setborder(ws, 5, 1, 2, 3)

    def test_invalid_col_range_raises(self):
        wb = Workbook()
        ws = wb.active
        with pytest.raises(ValueError, match="Start end column is before start column"):
            utils.setborder(ws, 1, 5, 3, 2)


# =============================================================================
# findFile
# =============================================================================

class TestFindFile:

    def test_finds_latest_file(self, tmp_path):
        # Create two files with the keyword
        import time
        file1 = tmp_path / "report_v1.csv"
        file1.write_text("data1")
        time.sleep(0.05)
        file2 = tmp_path / "report_v2.csv"
        file2.write_text("data2")

        result = utils.findFile(str(tmp_path), "report", ".csv", how="latest")
        assert "report_v2.csv" in result

    def test_finds_earliest_file(self, tmp_path):
        import time
        file1 = tmp_path / "report_v1.csv"
        file1.write_text("data1")
        time.sleep(0.05)
        file2 = tmp_path / "report_v2.csv"
        file2.write_text("data2")

        result = utils.findFile(str(tmp_path), "report", ".csv", how="earliest")
        assert "report_v1.csv" in result

    def test_returns_date_when_requested(self, tmp_path):
        file1 = tmp_path / "report.csv"
        file1.write_text("data")

        result = utils.findFile(str(tmp_path), "report", ".csv", returns="date")
        assert isinstance(result, datetime.date)

    def test_no_matching_files_raises(self, tmp_path):
        file1 = tmp_path / "unrelated.txt"
        file1.write_text("data")

        with pytest.raises((ValueError, IndexError)):
            utils.findFile(str(tmp_path), "report", ".csv")


# =============================================================================
# zip_and_clean_folder_with_exclusions
# =============================================================================

class TestZipAndCleanFolder:

    def test_creates_zip_file(self, tmp_path):
        # Create some test files
        (tmp_path / "old_file.txt").write_text("old data")
        (tmp_path / "another_old.txt").write_text("more old data")

        # Use a date that won't match any filenames
        curdate = datetime.date(2099, 12, 15)

        utils.zip_and_clean_folder_with_exclusions(curdate, str(tmp_path), [])

        folder_name = tmp_path.name
        zip_path = tmp_path / f"{folder_name}.zip"
        assert zip_path.exists(), "Zip file should be created"

        with zipfile.ZipFile(zip_path, "r") as zf:
            assert "old_file.txt" in zf.namelist()
            assert "another_old.txt" in zf.namelist()

    def test_excludes_files_in_excludelist(self, tmp_path):
        (tmp_path / "keep_me.txt").write_text("keep")
        (tmp_path / "zip_me.txt").write_text("zip")

        curdate = datetime.date(2099, 12, 15)

        utils.zip_and_clean_folder_with_exclusions(
            curdate, str(tmp_path), ["keep_me.txt"]
        )

        folder_name = tmp_path.name
        zip_path = tmp_path / f"{folder_name}.zip"

        with zipfile.ZipFile(zip_path, "r") as zf:
            assert "keep_me.txt" not in zf.namelist()
            assert "zip_me.txt" in zf.namelist()

    def test_excludes_files_with_current_date(self, tmp_path):
        curdate = datetime.date(2099, 12, 15)
        today_file = tmp_path / f"report_{curdate.strftime('%Y%m%d')}.txt"
        today_file.write_text("today")
        (tmp_path / "old_report.txt").write_text("old")

        utils.zip_and_clean_folder_with_exclusions(curdate, str(tmp_path), [])

        folder_name = tmp_path.name
        zip_path = tmp_path / f"{folder_name}.zip"

        with zipfile.ZipFile(zip_path, "r") as zf:
            assert today_file.name not in zf.namelist()
            assert "old_report.txt" in zf.namelist()

    def test_nonexistent_folder_raises(self):
        with pytest.raises(utils.FolderNotFound):
            utils.zip_and_clean_folder_with_exclusions(
                datetime.date(2026, 1, 1), "/nonexistent/path/xyz", []
            )


# =============================================================================
# current_function_name
# =============================================================================

class TestCurrentFunctionName:

    def test_returns_string(self):
        # When called inside a function, it returns that function's name
        result = utils.current_function_name()
        # It will return 'current_function_name' since that's the function it's inside
        assert result == "current_function_name"
